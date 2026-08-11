"""Modulo novo: consolidacao da base final (secao "Download da Base
Higienizada e Consolidacao dos Dados" do PDF "Automacao Astor Tech -> Nova
Vida", v2).

Junta o resultado baixado do Nova Vida de volta na base tratada do Astor
Tech, produzindo a base final pronta para a operacao.

CONFIRMADO com arquivo real (data/novavida/teste_resultado.zip ->
NV_UY3_..._LAYOUT.CSV, job de 116 CPFs baixado em 20/07/2026): a saida do
Nova Vida tem 56 colunas, separador ";", encoding latin-1 (acentos quebram
em UTF-8). As colunas usadas aqui batem exatamente com as letras Excel
citadas no PDF: Y=DDDCEL1, Z=CEL1, AB=FLWHATSAPPCEL1 (valores "S"/"N"/
vazio), AJ=EMAIL1.

Regras de negocio (confirmadas com o usuario nesta sessao):
  - So considera Telefone/E-mail de leads com FLWHATSAPPCEL1 == "S".
  - Telefone = DDDCEL1 + CEL1 (concatenados).
  - Merge e' LEFT (nao inner): leads sem WhatsApp/match ficam na base final
    mesmo assim, com Telefone/Email em branco - decisao do usuario de nao
    descartar leads elegiveis so por falta de contato.

ATUALIZADO em 04/08/2026 (pedido explicito do usuario): a coluna vinda de
settings.COLUNA_MARGEM_DISPONIVEL (campo real "available_margin" do Astor
Tech) e' rotulada "Parcela" na base final, seguindo a nomenclatura do PDF
original - mesmo o dado por tras sendo margem disponivel, nao um valor de
parcela de fato (o Astor Tech nao exporta valor de parcela real, ver
settings.py). A base final e' ordenada por Consulta decrescente (lead mais
recente primeiro) e salva em settings.DATA_ENTREGAS_DIR, separada dos
arquivos intermediarios de data/final/.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings
from src.logging_setup import get_logger

logger = get_logger("consolidacao")

_COR_CABECALHO = "1F4E78"
_COR_ZEBRA = "F2F6FA"
# Nomes cobrem tanto a base final consolidada (pos Nova Vida, colunas ja
# renomeadas) quanto a base tratada do Astor Tech pre-Nova Vida (nomes
# originais de data_treatment.py) - gerar_copia_formatada() e' usada nos
# dois casos (ver src/notificacao.py).
_COLUNAS_MOEDA = ("Valor Liberado", "Parcela", "Margem Disponivel")
_COLUNAS_DATA = ("Consulta", "Data da Consulta")
_COLUNAS_TEXTO = ("CPF", "Telefone")  # preserva zero a esquerda / nao vira numero
_LARGURA_MAXIMA_COLUNA = 42


def _ler_resultado_novavida(caminho: Path) -> pd.DataFrame:
    # index_col=False e' obrigatorio aqui: confirmado com arquivo real
    # (teste_resultado.zip) que cada linha de dados tem 1 campo a mais do
    # que o cabecalho (extra campo vazio no fim da linha). Sem
    # index_col=False, o pandas assume por padrao que a PRIMEIRA coluna e'
    # um indice sem nome, empurrando CPF/DDDCEL1/CEL1/FLWHATSAPPCEL1/EMAIL1
    # todos uma posicao para o lugar errado (bug confirmado: sem essa opcao,
    # FLWHATSAPPCEL1 aparecia com 0 valores "S" em vez dos 93/102 reais).
    df = pd.read_csv(caminho, sep=";", encoding="latin-1", dtype={"CPF": str}, index_col=False)
    df["CPF"] = df["CPF"].str.strip()

    com_whatsapp = df[df["FLWHATSAPPCEL1"].astype(str).str.strip().str.upper() == "S"].copy()
    com_whatsapp["Telefone"] = (
        com_whatsapp["DDDCEL1"].astype(str).str.strip() + com_whatsapp["CEL1"].astype(str).str.strip()
    )
    logger.info(
        "Resultado Nova Vida: %d linhas -> %d com WhatsApp disponivel (FLWHATSAPPCEL1=S)",
        len(df),
        len(com_whatsapp),
    )
    return com_whatsapp[["CPF", "Telefone", "EMAIL1"]].rename(columns={"EMAIL1": "Email"})


def consolidar(base_tratada: Path, resultado_novavida: Path) -> Path:
    tratada = pd.read_csv(base_tratada, dtype={"CPF": str})
    tratada["CPF"] = tratada["CPF"].str.strip()

    contatos = _ler_resultado_novavida(resultado_novavida)

    final = tratada.merge(contatos, on="CPF", how="left")
    final = final.rename(
        columns={
            "Numero de Parcelas": "Prazo",
            "Margem Disponivel": "Parcela",
            "Data da Consulta": "Consulta",
        }
    )
    final["Consulta"] = pd.to_datetime(final["Consulta"], errors="coerce")
    final = final.sort_values("Consulta", ascending=False)
    final = final[
        ["CPF", "Nome", "Valor Liberado", "Parcela", "Prazo", "Telefone", "Consulta", "Email"]
    ]

    sem_contato = final["Telefone"].isna().sum()
    logger.info(
        "Consolidacao concluida: %d leads na base final (%d sem Telefone/Email encontrado), "
        "ordenada por Consulta decrescente",
        len(final),
        sem_contato,
    )

    destino = settings.DATA_ENTREGAS_DIR / f"{base_tratada.stem}_consolidado.csv"
    final.to_csv(destino, index=False, encoding="utf-8-sig")
    logger.info("Base final consolidada salva em: %s", destino)

    gerar_copia_formatada(destino)
    return destino


def gerar_copia_formatada(caminho_csv: Path) -> Path:
    """Gera uma copia .xlsx da base final para leitura humana - pedido
    explicito do usuario em 04/08/2026 apos ver o CSV abrindo tudo
    espremido na coluna A no Excel (Excel pt-BR usa ";" como separador de
    lista, entao um CSV separado por "," nao quebra em colunas sozinho).

    O .xlsx resolve isso de raiz (nao depende de separador nenhum) e ainda
    aplica formatacao profissional: cabecalho destacado, zebra listrado,
    moeda em Valor Liberado/Parcela, CPF/Telefone como texto (preserva
    zero a esquerda), Consulta como data/hora legivel, colunas
    autoajustadas, filtro automatico e painel congelado no cabecalho.
    Ordenado por Consulta decrescente (lead mais recente primeiro).

    ATUALIZADO em 05/08/2026 (pedido explicito do usuario): o filtro
    automatico da coluna Telefone ja vem aplicado com "Vazio" desmarcado -
    ao abrir o arquivo, leads sem Telefone ja ficam ocultos, sem precisar
    o usuario clicar em nada. Isso exige duas coisas (uma so nao basta):
    1) gravar a definicao do filtro (FilterColumn) com a lista de valores
    marcados, sem incluir blank; 2) esconder de fato as linhas em branco
    via row_dimensions[].hidden - o Excel confia no estado salvo no
    arquivo e so recalcula quando o usuario reabre o dropdown do filtro.
    """
    df = pd.read_csv(caminho_csv, dtype={col: str for col in _COLUNAS_TEXTO})
    coluna_data = next((c for c in _COLUNAS_DATA if c in df.columns), None)
    if coluna_data:
        df[coluna_data] = pd.to_datetime(df[coluna_data], errors="coerce")
        df = df.sort_values(coluna_data, ascending=False)

    destino = caminho_csv.with_suffix(".xlsx")
    df.to_excel(destino, index=False, sheet_name="Leads")

    wb = load_workbook(destino)
    ws = wb["Leads"]

    borda = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    total_linhas, total_colunas = ws.max_row, ws.max_column

    for col_idx in range(1, total_colunas + 1):
        celula_cabecalho = ws.cell(row=1, column=col_idx)
        celula_cabecalho.font = Font(color="FFFFFF", bold=True, size=11)
        celula_cabecalho.fill = PatternFill("solid", fgColor=_COR_CABECALHO)
        celula_cabecalho.alignment = Alignment(horizontal="center", vertical="center")
        celula_cabecalho.border = borda

        nome_coluna = df.columns[col_idx - 1]
        maior_valor = max(
            [len(str(nome_coluna))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            maior_valor + 3, _LARGURA_MAXIMA_COLUNA
        )

        for linha_idx in range(2, total_linhas + 1):
            celula = ws.cell(row=linha_idx, column=col_idx)
            celula.border = borda
            if linha_idx % 2 == 0:
                celula.fill = PatternFill("solid", fgColor=_COR_ZEBRA)
            if nome_coluna in _COLUNAS_MOEDA:
                celula.number_format = "R$ #,##0.00"
                celula.alignment = Alignment(horizontal="right")
            elif nome_coluna in _COLUNAS_DATA:
                celula.number_format = "dd/mm/yyyy hh:mm"
                celula.alignment = Alignment(horizontal="center")
            elif nome_coluna in _COLUNAS_TEXTO:
                celula.number_format = "@"
            elif nome_coluna == "Prazo":
                celula.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    ws.auto_filter.ref = f"A1:{get_column_letter(total_colunas)}{total_linhas}"

    if "Telefone" in df.columns:
        telefone_vazio = df["Telefone"].isna() | (df["Telefone"].astype(str).str.strip() == "")
        col_id_telefone = df.columns.get_loc("Telefone")
        valores_com_telefone = sorted(df.loc[~telefone_vazio, "Telefone"].astype(str).unique().tolist())
        ws.auto_filter.add_filter_column(col_id_telefone, valores_com_telefone, blank=False)
        for offset, vazio in enumerate(telefone_vazio):
            if vazio:
                ws.row_dimensions[offset + 2].hidden = True
        logger.info(
            "Filtro de Telefone pre-aplicado: %d leads sem telefone ocultos ao abrir (de %d)",
            int(telefone_vazio.sum()),
            len(df),
        )

    wb.save(destino)
    logger.info("Copia formatada (.xlsx) salva em: %s (%d leads)", destino, total_linhas - 1)
    return destino


if __name__ == "__main__":
    import sys

    consolidar(Path(sys.argv[1]), Path(sys.argv[2]))
